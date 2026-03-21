import openpyxl

class ReadExcel:
    def __init__(self,file_path):
        self.workbook = openpyxl.load_workbook(file_path)

    def get_data(self,sheet_name):
        sheet = self.workbook[sheet_name]
        data = []

        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)  # skip header

        for row in rows:
            data.append(dict(zip(headers, row)))

        return data

    # def get_valid_login(self):
    #     sheet = self.workbook.active
    #     rows = sheet.iter_rows(values_only=True)
    #     headers = next(rows)  # skip
    #     valid = next(rows)
    #     return [valid[1],valid[2]]

    def close(self):
        self.workbook.close()



# file = ReadExcel(r"C:\vani\Selenium\PycharmProjects\files\saucedemo_login.xlsx")
# print(file.get_valid_login)